import copy

from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border, Style, NumberFormatDescriptor, Side


from openpyxl.cell import Cell
import datetime

colorMap = {'Chargeback': 'FF6600', 'Terminated': 'FF00FF', 'Active': '00FFFFFF'}
'0.00E+00'

cell_number_format = {'number': '0.00E+00', 'currency': '"$"#,##0.00_-', 'percentage': '00.00%',
                      'date': 'dd/mm/yy'}


def create_commission_agent(sales, payment_date, branch, shift):
    wb = Workbook()
    # ws = wb.active
    # ws_sumary = wb.create_sheet()
    # ws_sumary.title = "Earnings Summary"
    #
    # addcell_value(1, 1, 5, '00FF0000', ws_sumary)
    # addcell_value(2, 1, 6, '00CCFFCC', ws_sumary)
    # addcell_value(3, 1, "=SUM(A1, A2)", '00FFFFFF', ws_sumary, '0.00%')
    # addcell_value(5, 1, payment_date, '00CCFFCC', ws_sumary, data_type='mm/dd/yyyy')
    d_file = payment_date

    total_tmpl = {'total_sales': 0,
                  'total_charge_back': 0,
                  'total_terminated': 0,
                  'amount_sales': 0,
                  'amount_charge_back': 0,
                  'amount_terminated': 0,
                  'is_core': False,
                  'carrier': ''}
    summary = {}
    formula_total_written = []
    formula_total_terminated = []
    formula_total_chargeback = []
    push_terms = {}

    for key, value in sales.iteritems():
        counter = 3
        ws_agent_sheet = wb.create_sheet()
        agent_name = value['agent_name']
        ws_agent_sheet.title = agent_name
        add_header(branch, counter, d_file, shift, ws_agent_sheet)
        counter += 1
        first_row = 3
        total_status = {'Chargeback': """SUMIFS(M{}:M{}, I{}:I{}, "Chargeback")""",
                        'Terminated': """SUMIFS(M{}:M{}, I{}:I{}, "Terminated")""",
                        'Active': """SUMIFS(M{}:M{}, I{}:I{}, "Active")"""
                        }

        sales_detail = value['sales']
        for item in sales_detail:

            group = item['group']
            policy_status = item['status']
            carrier = item['carrier']
            is_core = item['is_core']
            earning_value = item['earning']
            add_row(agent_name, counter, earning_value, item, policy_status, ws_agent_sheet)

            prod_summary = summary.get(group)
            if prod_summary:
                fill_summary(carrier, earning_value, is_core, policy_status, prod_summary)
            else:
                prod_summary = copy.deepcopy(total_tmpl)
                fill_summary(carrier, earning_value, is_core, policy_status, prod_summary)
            counter += 1
            summary[group] = prod_summary

        last_row = counter + 2
        counter += 2
        addcell_value(counter, 2, agent_name, colorMap['Active'], ws_agent_sheet, number_format=None)
        addcell_value(counter, 3, d_file.strftime('%m.%d.%Y'), colorMap['Active'], ws_agent_sheet, number_format=None)

        total_active = total_status['Active'].format(first_row, last_row, first_row, last_row)

        total_terminated = total_status['Terminated'].format(first_row, last_row, first_row, last_row)
        total_charge_back = total_status['Chargeback'].format(first_row, last_row, first_row, last_row)
        add_footer(agent_name, counter, total_active, total_charge_back, total_terminated, value, ws_agent_sheet)
        tmpl_summary_totals = """'{}'!C{}"""
        formula_total_written.append(tmpl_summary_totals.format(agent_name, str(counter + 1)))
        formula_total_terminated.append(tmpl_summary_totals.format(agent_name, str(counter + 3)))
        formula_total_chargeback.append(tmpl_summary_totals.format(agent_name, str(counter + 4)))
        push_terms[agent_name] = {'total_pay': 'C{}'.format(counter + 19),
                                  'new_push': 'C{}'.format(counter + 18),
                                  }

    if summary:
        ws_earning_summary = wb.create_sheet()
        ws_earning_summary.title = 'Earnings Summary'
        create_summary(branch, summary, ws_earning_summary, payment_date, payment_date, formula_total_written,
                       formula_total_chargeback, formula_total_terminated)

    filename = '../{}_{}.xlsx'.format(branch, d_file.strftime('%m_%d_%Y'))

    ws_push_terms = wb.create_sheet()
    ws_push_terms.title = 'Push-Terms'
    create_push_terms(ws_push_terms, push_terms)
    wb.save(filename)


def add_footer(agent_name, counter, total_active, total_charge_back, total_terminated, value, ws_agent_sheet):
    written = """{} + {} """.format(total_active, total_terminated)

    addcell_value(counter + 1, 2, 'Written', colorMap['Active'], ws_agent_sheet, is_bold=True, number_format=None)
    addcell_value(counter + 1, 3, '={}'.format(written), colorMap['Active'], ws_agent_sheet,
                  number_format=cell_number_format['currency'])
    addcell_value(counter + 2, 2, 'Active', colorMap['Active'], ws_agent_sheet, is_bold=True, number_format=None)
    addcell_value(counter + 2, 3, '={}'.format(total_active), colorMap['Active'], ws_agent_sheet,
                  number_format=cell_number_format['currency'])
    addcell_value(counter + 3, 2, 'Terminated', colorMap['Terminated'], ws_agent_sheet, is_bold=True,
                  number_format=None)
    addcell_value(counter + 3, 3, '={}'.format(total_terminated), colorMap['Active'], ws_agent_sheet,
                  number_format=cell_number_format['currency'])
    addcell_value(counter + 4, 2, 'Chargeback', colorMap['Chargeback'], ws_agent_sheet, is_bold=True,
                  number_format=None)
    addcell_value(counter + 4, 3, '={}'.format(total_charge_back), colorMap['Active'], ws_agent_sheet,
                  number_format=cell_number_format['currency'])
    addcell_value(counter + 5, 2, 'Advances', colorMap['Active'], ws_agent_sheet, is_bold=True,
                  number_format=cell_number_format['currency'])
    total_paid = "=SUM(C{}, C{})".format(counter + 2, counter + 4)
    addcell_value(counter + 6, 2, 'Total Pay this Period', colorMap['Active'], ws_agent_sheet, is_bold=True,
                  number_format=None)
    addcell_value(counter + 6, 3, total_paid, colorMap['Active'], ws_agent_sheet,
                  number_format=cell_number_format['currency'])
    percentage = "=(C{})/C{}".format(counter + 6, counter + 1)
    addcell_value(counter + 10, 2, 'Percentage', colorMap['Active'], ws_agent_sheet, is_bold=True, number_format=None)
    addcell_value(counter + 10, 3, percentage, colorMap['Active'], ws_agent_sheet,
                  number_format=cell_number_format['percentage'])
    addcell_value(counter + 12, 2, 'Push/Terms from ', colorMap['Active'], ws_agent_sheet, is_bold=True,
                  number_format=None)
    addcell_value(counter + 12, 3, '', colorMap['Active'], ws_agent_sheet, number_format=cell_number_format['currency'])
    addcell_value(counter + 13, 2, 'Push/Terms Adj.', colorMap['Active'], ws_agent_sheet, is_bold=True,
                  number_format=None)
    addcell_value(counter + 13, 3, '', colorMap['Active'], ws_agent_sheet, number_format=cell_number_format['currency'])
    addcell_value(counter + 14, 2, 'New Total:', colorMap['Active'], ws_agent_sheet, is_bold=True, number_format=None)
    addcell_value(counter + 14, 3, "=C{}-C{}".format(counter + 6, counter + 12), colorMap['Active'], ws_agent_sheet,
                  number_format=cell_number_format['currency'])
    payable = value['payable'] if value['payable'] else agent_name
    addcell_value(counter + 15, 2, "Payable To:", colorMap['Active'], ws_agent_sheet, is_bold=True, number_format=None)
    addcell_value(counter + 15, 3, payable, colorMap['Active'], ws_agent_sheet,
                  number_format=cell_number_format['currency'])
    addcell_value(counter + 17, 2, 'New Total:', colorMap['Active'], ws_agent_sheet, is_bold=True, number_format=None)
    addcell_value(counter + 17, 3, '=C{}'.format(counter + 14), colorMap['Active'], ws_agent_sheet,
                  number_format=cell_number_format['currency'])
    addcell_value(counter + 18, 2, 'New Push/Terms', colorMap['Active'], ws_agent_sheet, is_bold=True,
                  number_format=None)
    addcell_value(counter + 18, 3, '=SUM(C{}:C{})'.format(counter + 12, counter + 13), colorMap['Active'],
                  ws_agent_sheet, number_format=cell_number_format['currency'])
    addcell_value(counter + 19, 2, 'Total Pay:', colorMap['Active'], ws_agent_sheet, is_bold=True, number_format=None)
    addcell_value(counter + 19, 3, '=SUM(C{}:C{})'.format(counter + 17, counter + 16), colorMap['Active'],
                  ws_agent_sheet, number_format=cell_number_format['currency'])
    addcell_value(counter + 19, 4, '=(C{})/C{}'.format(counter + 18, counter + 1), colorMap['Active'],
                  ws_agent_sheet, number_format=cell_number_format['percentage'])


def add_row(agent_name, counter, earning_value, item, policy_status, ws_agent_sheet):
    addcell_value(counter, 1, item['counter'], colorMap[policy_status], ws_agent_sheet, number_format=None)
    addcell_value(counter, 2, item['member_id'], colorMap[policy_status], ws_agent_sheet, number_format=None)
    addcell_value(counter, 3, item['group'], colorMap[policy_status], ws_agent_sheet, number_format=None)
    addcell_value(counter, 4, agent_name, colorMap[policy_status], ws_agent_sheet, number_format=None)
    addcell_value(counter, 5, item['first_name'], colorMap[policy_status], ws_agent_sheet, number_format=None)
    addcell_value(counter, 6, item['last_name'], colorMap[policy_status], ws_agent_sheet, number_format=None)
    addcell_value(counter, 7, item['effective_date'], colorMap[policy_status], ws_agent_sheet,
                  number_format=cell_number_format['date'])
    addcell_value(counter, 8, item['enrollment_date'], colorMap[policy_status], ws_agent_sheet,
                  number_format=cell_number_format['date'])
    addcell_value(counter, 9, item['status'], colorMap[policy_status], ws_agent_sheet, number_format=None)
    addcell_value(counter, 10, item['cancel_date'], colorMap[policy_status], ws_agent_sheet,
                  number_format=cell_number_format['date'])
    addcell_value(counter, 11, item['plan_type'], colorMap[policy_status], ws_agent_sheet, number_format=None)
    addcell_value(counter, 12, item['coverage_type'], colorMap[policy_status], ws_agent_sheet, number_format=None)
    earning = earning_value if policy_status in ['Active', 'Terminated'] else earning_value * -1
    addcell_value(counter, 13, earning, colorMap[policy_status], ws_agent_sheet,
                  number_format=cell_number_format['currency'])


def add_header(branch, counter, d_file, shift, ws_agent_sheet):
    addcell_value(1, 3, 'Profile', colorMap['Active'], ws_agent_sheet, is_bold=True)
    addcell_value(1, 4, branch, colorMap['Active'], ws_agent_sheet, is_bold=True)
    addcell_value(1, 5, 'Shift', colorMap['Active'], ws_agent_sheet, is_bold=True)
    addcell_value(1, 6, shift, colorMap['Active'], ws_agent_sheet, is_bold=True)
    addcell_value(1, 8, 'Created:{}'.format(d_file.strftime('%m.%d.%Y')), colorMap['Active'], ws_agent_sheet,
                  is_bold=True)
    addcell_value(1, 10, 'Printed:{}'.format(d_file.strftime('%m.%d.%Y')), colorMap['Active'], ws_agent_sheet,
                  is_bold=True)
    addcell_value(1, 12, 'PayDate:{}'.format(d_file.strftime('%m.%d.%Y')), colorMap['Active'], ws_agent_sheet,
                  is_bold=True)
    addcell_value(counter, 1, 'Counter', colorMap['Active'], ws_agent_sheet, is_bold=True)
    addcell_value(counter, 2, 'Member Id', colorMap['Active'], ws_agent_sheet, is_bold=True)
    addcell_value(counter, 3, 'Group', colorMap['Active'], ws_agent_sheet, is_bold=True)
    addcell_value(counter, 4, 'Fronter', colorMap['Active'], ws_agent_sheet, is_bold=True)
    addcell_value(counter, 5, 'First Name', colorMap['Active'], ws_agent_sheet, is_bold=True)
    addcell_value(counter, 6, 'Last Name', colorMap['Active'], ws_agent_sheet, is_bold=True)
    addcell_value(counter, 7, 'Effective Date', colorMap['Active'], ws_agent_sheet, is_bold=True)
    addcell_value(counter, 8, 'Enrollment Date', colorMap['Active'], ws_agent_sheet, is_bold=True)
    addcell_value(counter, 9, 'Status', colorMap['Active'], ws_agent_sheet, is_bold=True)
    addcell_value(counter, 10, 'Cancel Date', colorMap['Active'], ws_agent_sheet, is_bold=True)
    addcell_value(counter, 11, 'Plan', colorMap['Active'], ws_agent_sheet, is_bold=True)
    addcell_value(counter, 12, 'Coverage Type', colorMap['Active'], ws_agent_sheet, is_bold=True)
    addcell_value(counter, 13, 'Earning', colorMap['Active'], ws_agent_sheet, is_bold=True)


def create_summary(branch, summary, ws_earning_summary, pay_date, printed_date, formula_total_written,
                   formula_total_chargeback, formula_total_terminated):
    addcell_value(1, 4, 'Profile:'.format(branch), colorMap['Active'], ws_earning_summary, is_bold=True,
                  number_format=None)
    addcell_value(2, 4, 'PayDate_date:{}'.format(pay_date.strftime('%m.%d.%Y')), colorMap['Active'], ws_earning_summary,
                  is_bold=True, number_format=None)
    addcell_value(2, 6, 'Earnings Summary', colorMap['Active'], ws_earning_summary, is_bold=True,
                  number_format=None)
    addcell_value(3, 4, 'Printed:{}'.format(printed_date.strftime('%m.%d.%Y')), colorMap['Active'], ws_earning_summary,
                  is_bold=True, number_format=None)
    addcell_value(4, 7, 'Dollar Ammount', colorMap['Active'], ws_earning_summary, is_bold=True, number_format=None)
    addcell_value(4, 8, 'Percentage', colorMap['Active'], ws_earning_summary, is_bold=True, number_format=None)
    terminated = 0
    charge_back = 0
    sales = 0
    core_total = {'sales_amount': 0,
                  'sales_total': 0,
                  'charge_back_amount': 0,
                  'charge_back_total': 0,
                  'terminated_amount': 0,
                  'terminated_total': 0
                  }
    ancillary_total = {'sales_amount': 0,
                       'sales_total': 0,
                       'charge_back_amount': 0,
                       'charge_back_total': 0,
                       'terminated_amount': 0,
                       'terminated_total': 0
                       }

    charge_back = 0
    sales = 0
    count_product = len(summary.keys())

    row_total = 25

    addcell_value(row_total, 6, 'Group', colorMap['Active'], ws_earning_summary, is_bold=True)
    addcell_value(row_total, 7, 'Total Count', colorMap['Active'], ws_earning_summary, is_bold=True)

    for key, value in summary.iteritems():
        # terminated += value.get('amount_terminated')
        # charge_back += value.get('amount_charge_back')
        # sales += value.get('amount_sales')
        if value.get('is_core'):
            fill_core_ancillary(core_total, value)
        else:
            fill_core_ancillary(ancillary_total, value)
        row_total += 1
        carrier = value.get('carrier')
        addcell_value(row_total, 6, 'Sales {} ({})'.format(key, carrier), colorMap['Active'], ws_earning_summary,
                      is_bold=True)
        addcell_value(row_total, 7, '{}'.format(value.get('total_sales')), colorMap['Active'], ws_earning_summary,
                      number_format=cell_number_format['currency'])

        addcell_value(row_total + count_product, 6, 'Terminated {} ({})'.format(key, carrier), colorMap['Active'],
                      ws_earning_summary, is_bold=True)
        addcell_value(row_total + count_product, 7, '{}'.format(value.get('total_terminated')), colorMap['Active'],
                      ws_earning_summary, number_format=cell_number_format['currency'])

        addcell_value(row_total + 2 * count_product, 6, 'Charge back {} ({})'.format(key, carrier), colorMap['Active'],
                      ws_earning_summary, is_bold=True)
        addcell_value(row_total + 2 * count_product, 7, '{}'.format(value.get('total_charge_back')), colorMap['Active'],
                      ws_earning_summary, number_format=cell_number_format['currency'])

    terminated = '+'.join(formula_total_terminated)
    written = '+'.join(formula_total_written)
    charge_back = '+'.join(formula_total_chargeback)

    addcell_value(5, 6, 'Written', colorMap['Active'], ws_earning_summary, is_bold=True)
    addcell_value(5, 7, '={}'.format(written), colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['currency'])
    addcell_value(5, 8, '100', colorMap['Active'], ws_earning_summary, number_format=cell_number_format['percentage'])

    addcell_value(6, 6, 'Charge back', colorMap['Chargeback'], ws_earning_summary, is_bold=True)
    addcell_value(6, 7, '={}'.format(charge_back), colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['currency'])
    addcell_value(6, 8, '=(-1*G6)/G5', colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['percentage'])

    addcell_value(7, 6, 'Terminated', colorMap['Terminated'], ws_earning_summary, is_bold=True)
    addcell_value(7, 7, '=({})*-1'.format(terminated), colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['currency'])
    addcell_value(7, 8, '=(-G7)/G5', colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['percentage'])

    addcell_value(8, 6, 'Pushes', '23fcec', ws_earning_summary, is_bold=True)
    addcell_value(9, 6, 'Prior Period Push', '23fcec', ws_earning_summary, is_bold=True)
    addcell_value(10, 6, 'Terms', '23fcec', ws_earning_summary, is_bold=True)

    # total_period = sales - charge_back
    total_period = 1
    addcell_value(11, 6, 'Total Pay this Period', colorMap['Active'], ws_earning_summary, is_bold=True)
    addcell_value(11, 7, '=SUM(G5:G7)', colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['currency'])
    addcell_value(11, 8, '=(G11/G5)', colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['percentage'])

    addcell_value(15, 7, 'Dollar Ammount', colorMap['Active'], ws_earning_summary, is_bold=True)
    addcell_value(15, 8, 'Percentage', colorMap['Active'], ws_earning_summary, is_bold=True,
                  number_format=cell_number_format['currency'])
    addcell_value(15, 9, 'Total Count', colorMap['Active'], ws_earning_summary, is_bold=True,
                  number_format=cell_number_format['percentage'])

    total_amount = core_total.get('sales_amount')
    addcell_value(16, 6, 'Total Sales CORE', '2333fc', ws_earning_summary, is_bold=True)
    addcell_value(16, 7, total_amount, colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['currency'])
    addcell_value(16, 8, '=(G16)/G11', colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['percentage'])
    addcell_value(16, 9, core_total.get('sales_total'), colorMap['Active'], ws_earning_summary)

    total_amount = ancillary_total.get('sales_amount')
    addcell_value(17, 6, 'Total Sales Ancillary', '2333fc', ws_earning_summary, is_bold=True)
    addcell_value(17, 7, total_amount, colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['currency'])
    addcell_value(17, 8, '=(G17)/G11', colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['percentage'])
    addcell_value(17, 9, ancillary_total.get('sales_total'), colorMap['Active'], ws_earning_summary)

    total_amount = core_total.get('terminated_amount')
    addcell_value(18, 6, 'Total Termination CORE', colorMap['Terminated'], ws_earning_summary, is_bold=True)
    addcell_value(18, 7, total_amount, colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['currency'])
    addcell_value(18, 8, '=(G18)/G11', colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['percentage'])
    addcell_value(18, 9, core_total.get('terminated_total'), colorMap['Active'], ws_earning_summary)

    total_amount = ancillary_total.get('terminated_amount')
    addcell_value(19, 6, 'Total Termination Ancillary', colorMap['Terminated'], ws_earning_summary, is_bold=True)
    addcell_value(19, 7, total_amount, colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['currency'])
    addcell_value(19, 8, '=(G19)/G11', colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['percentage'])
    addcell_value(19, 9, ancillary_total.get('terminated_total'), colorMap['Active'], ws_earning_summary)

    total_amount = core_total.get('charge_back_amount')
    addcell_value(20, 6, 'Total Charge back CORE', colorMap['Chargeback'], ws_earning_summary, is_bold=True)
    addcell_value(20, 7, total_amount, colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['currency'])
    addcell_value(20, 8, '=(G20)/G11', colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['percentage'])
    addcell_value(20, 9, core_total.get('charge_back_total'), colorMap['Active'], ws_earning_summary)

    total_amount = ancillary_total.get('charge_back_amount')
    addcell_value(21, 6, 'Total Charge back Ancillary', colorMap['Chargeback'], ws_earning_summary, is_bold=True)
    addcell_value(21, 7, total_amount, colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['currency'])
    addcell_value(21, 8, '=(G21)/G11', colorMap['Active'], ws_earning_summary,
                  number_format=cell_number_format['percentage'])
    addcell_value(21, 9, ancillary_total.get('charge_back_total'), colorMap['Active'], ws_earning_summary)


def create_push_terms(ws_push_terms, push_terms):
    row = 1
    addcell_value(row, 1, 'Agent', colorMap['Active'], ws_push_terms)
    addcell_value(row, 2, 'New Push/Terms', colorMap['Active'], ws_push_terms)
    addcell_value(row, 3, 'Total Pay', colorMap['Active'], ws_push_terms)
    tmpl = """='{}'!{}"""
    for key, value in push_terms.iteritems():
        row += 1

        addcell_value(row, 1, key, colorMap['Active'], ws_push_terms)
        addcell_value(row, 2, tmpl.format(key, value['new_push']), colorMap['Active'], ws_push_terms,
                      number_format=cell_number_format['currency'])
        addcell_value(row, 3, tmpl.format(key, value['total_pay']), colorMap['Active'], ws_push_terms,
                      number_format=cell_number_format['currency'])


def fill_core_ancillary(summary_dict, value):
    summary_dict['sales_amount'] += value.get('amount_sales')
    summary_dict['charge_back_amount'] += value.get('amount_charge_back')
    summary_dict['terminated_amount'] += value.get('amount_terminated')
    summary_dict['sales_total'] += value.get('total_sales')
    summary_dict['charge_back_total'] += value.get('total_charge_back')
    summary_dict['terminated_total'] += value.get('total_terminated')


def fill_summary(carrier, earning_value, is_core, policy_status, prod_summary):
    if 'Terminated' in policy_status:
        prod_summary['amount_terminated'] += earning_value
        prod_summary['total_terminated'] += 1
    elif 'Active' in policy_status:
        prod_summary['amount_sales'] += earning_value
        prod_summary['total_sales'] += 1
    else:
        prod_summary['amount_charge_back'] += earning_value
        prod_summary['total_charge_back'] += 1
    prod_summary['carrier'] = carrier
    prod_summary['is_core'] = is_core


def addcell_value(row, col, data, bg_color, sheet, number_format=None, font_size=12, is_bold=False, ):
    cell = sheet.cell(row=row, column=col)
    column = cell.column
    cell = '{}{}'.format(cell.column, cell.row)
    sheet[cell] = data

    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="thin", color="000000")
    border = Border(top=double, left=thin, right=thin, bottom=double)

    sheet[cell].style = Style(fill=PatternFill(patternType='solid', fgColor=Color(bg_color)), border=border,
                              number_format=number_format, font=Font(size=font_size, bold=is_bold))
    if '=' not in str(data):
        if not sheet.column_dimensions[column].width or sheet.column_dimensions[column].width < len(str(data)):
            sheet.column_dimensions[column].width = len(str(data))


def main():
    tmp = {'counter': 0,
           'member_id': '0',
           'group': '',
           'first_name': '',
           'last_name': '',
           'effective_date': '',
           'enrollment_date': '',
           'status': '',
           'cancel_date': '',
           'coverage_type': '',
           'earning': ''
           }
    s1 = copy.deepcopy(tmp)
    s1['member_id'] = 'CL13358500'
    s1['group'] = 'Principle Advantage'
    s1['first_name'] = 'Jerry'
    s1['last_name'] = 'Lewis'
    s1['effective_date'] = '8/23/2016'
    s1['enrollment_date'] = '8/22/2016'
    s1['status'] = 'Active'
    s1['cancel_date'] = ''
    s1['plan_type'] = 'Guaranteed Issue'
    s1['coverage_type'] = ' Individual'
    s1['earning'] = 35.80
    s1['carrier'] = 'HII'
    s1['is_core'] = True

    s2 = copy.deepcopy(tmp)
    s2['member_id'] = 'CL13389600'
    s2['group'] = 'Principle Advantage'
    s2['first_name'] = 'Jamie'
    s2['last_name'] = 'Oliver'
    s2['effective_date'] = '8/24/2016'
    s2['enrollment_date'] = '8/23/2016'
    s2['status'] = 'Terminated'
    s2['cancel_date'] = '8/25/2016'
    s2['plan_type'] = 'Guaranteed Issue'
    s2['coverage_type'] = '  Individual & Spouse'
    s2['earning'] = 45.20
    s2['carrier'] = 'HII'
    s2['is_core'] = True

    s3 = copy.deepcopy(tmp)
    s3['member_id'] = 'CL13389600'
    s3['group'] = 'Principle Advantage'
    s3['first_name'] = 'John'
    s3['last_name'] = 'Smith'
    s3['effective_date'] = '8/24/2016'
    s3['enrollment_date'] = '8/23/2016'
    s3['status'] = 'Chargeback'
    s3['cancel_date'] = '8/25/2016'
    s3['plan_type'] = 'Guaranteed Issue'
    s3['coverage_type'] = '  Individual & Spouse'
    s3['earning'] = 10.65
    s3['carrier'] = 'HII'
    s3['is_core'] = True

    s5 = copy.deepcopy(tmp)
    s5['member_id'] = 'PL13358500'
    s5['group'] = 'Freddom Spirit Plus'
    s5['first_name'] = 'Jerry'
    s5['last_name'] = 'Lewis'
    s5['effective_date'] = '8/23/2016'
    s5['enrollment_date'] = '8/22/2016'
    s5['status'] = 'Active'
    s5['cancel_date'] = ''
    s5['plan_type'] = 'Guaranteed Issue'
    s5['coverage_type'] = ' Individual'
    s5['earning'] = 35.05
    s5['carrier'] = 'HII'
    s5['is_core'] = False

    adamb = [s1, s2, s3, s5]

    sales = {'adamb': {'sales': adamb,
                       'agent_name': 'Adamb Berco',
                       'payable': 'Adamb Berco',

                       }}
    create_commission_agent(sales=sales, payment_date=datetime.date.today(), branch='HBC', shift='day')


if __name__ == '__main__':
    main()
